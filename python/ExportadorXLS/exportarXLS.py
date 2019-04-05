#!/usr/bin/python3

#from PyQt4 import QtCore, QtGui, QtSql
from openpyxl import Workbook
import time
#import xlwt

def exportar(ruta, *datos):
	print ("Ruta: " + ruta)
	for dato in datos:
		print("Datos: " + str(dato))
	'''book = Workbook()
	sheet = book.create_sheet(datos[0]) # insert at the end (default)
	a = sheet.cell(row=1, column = 1, value = datos[1])
	b = sheet.cell(row=1, column = 2, value = datos[2])
	c = sheet.cell (row=1, column = 3)
	cadena = '=' + a.column+str(a.row)+"*"+b.column+str(b.row)
	c.value = cadena
	print (cadena)
	print (a.column+str(a.row))	
	book.save(ruta)'''
		
#----------------------------------------------------------------------

if __name__ == '__main__':
	exportar("/home/david/kk.xlsx","Datos",3,4)    

    
