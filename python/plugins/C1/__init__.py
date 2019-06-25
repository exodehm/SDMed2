from PyQt5 import QtSql
from openpyxl import Workbook
from openpyxl.drawing import line
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Font
from enum import Enum

import datetime

class Columnas(Enum):
    CCodigo = 1
    CCantidad = 2
    CUd = 3
    CResumen = 4
    CPrecio = 5
    CImporte = 6

def run():
    print("Plugin para imprimir precios auxiliares de mano de obra, materiales y maquinaria")
    
def imprimir(conexion, obra, book):
	tipo = 1
	consulta = QtSql.QSqlQuery("SELECT * FROM ver_conceptos_cantidad('"+ obra+"','"+str(tipo)+"')", conexion)
	print (consulta.lastError().text())
	rec = consulta.record()
	codigo = rec.indexOf("codigo")
	cantidad = rec.indexOf("cantidad")
	ud = rec.indexOf("ud")
	resumen = rec.indexOf("resumen");
	precio = rec.indexOf("precio");
	#abro una instancia de hoja de calculo		
	sheet = book.active
	sheet.title = "Listado de Auxiliares"
	#fuentes
	ft_resaltada = Font(name='Arial', size=11, bold=True)
	ft_normal = Font(name='Arial', size=10, bold=False)
	ft_chica = Font(name='Arial', size=9, bold=False)	
	#cabecera
	sheet.oddHeader.left.size = 11
	sheet.oddHeader.left.font = "Arial,Bold"	
	sheet.oddHeader.left.text = "LISTADO DE OBRAS VALORADO "	
	sheet.print_title_cols = 'A:C'
	sheet.print_title_rows = '1:2'	
	#pie
	sheet.oddFooter.center.size = 11
	sheet.oddFooter.center.font = "Arial,Bold"	
	sheet.oddFooter.center.text = "Página &[Page] de &N"	
	#e inserto datos
	#datos para situar filas y columnas
	fila = 1
	#print (consulta.size())
	#cabecera
	sheet.cell(column = Columnas.CCodigo.value, row = fila).value = "CODIGO"
	sheet.cell(column = Columnas.CCodigo.value, row = fila).font = ft_normal
	sheet.cell(column = Columnas.CCantidad.value, row = fila).value = "CANTIDAD"
	sheet.cell(column = Columnas.CCantidad.value, row = fila).font = ft_normal
	sheet.cell(column = Columnas.CCantidad.value, row = fila).alignment = Alignment(horizontal="right")
	sheet.cell(column = Columnas.CUd.value, row = fila).value = "UD"
	sheet.cell(column = Columnas.CUd.value, row = fila).font = ft_normal
	sheet.cell(column = Columnas.CUd.value, row = fila).alignment = Alignment(horizontal="right")
	sheet.cell(column = Columnas.CResumen.value, row = fila).value = "RESUMEN"
	sheet.cell(column = Columnas.CResumen.value, row = fila).font = ft_normal
	sheet.cell(column = Columnas.CPrecio.value, row = fila).value = "PRECIO"
	sheet.cell(column = Columnas.CPrecio.value, row = fila).font = ft_normal
	sheet.cell(column = Columnas.CPrecio.value, row = fila).alignment = Alignment(horizontal="right", vertical="center")
	sheet.cell(column = Columnas.CImporte.value, row = fila).value = "IMPORTE"
	sheet.cell(column = Columnas.CImporte.value, row = fila).font = ft_normal
	sheet.cell(column = Columnas.CImporte.value, row = fila).alignment = Alignment(horizontal="right", vertical="center")
	#linea cabecera
	for i in range(Columnas.CCodigo.value,Columnas.CImporte.value+1):
		sheet.cell(column = i, row = fila).border = Border(bottom=Side(style='thick'))		
	#datos
	fila = fila +1
	while consulta.next():		
		#codigo
		sheet.cell(column = Columnas.CCodigo.value, row = fila).value = consulta.value(codigo)
		sheet.cell(column = Columnas.CCodigo.value, row = fila).font = ft_normal		
		#cantidad
		sheet.cell(column = Columnas.CCantidad.value, row = fila).value = consulta.value(cantidad)
		sheet.cell(column = Columnas.CCantidad.value, row = fila).font = ft_normal
		sheet.cell(column = Columnas.CCantidad.value, row = fila).number_format = '#,##0.00'
		#ud
		sheet.cell(column = Columnas.CUd.value, row = fila).value = consulta.value(ud)
		sheet.cell(column = Columnas.CUd.value, row = fila).font = ft_normal
		sheet.cell(column = Columnas.CUd.value, row = fila).alignment = Alignment(horizontal="right")
		#resumen
		sheet.cell(column = Columnas.CResumen.value, row = fila).value = consulta.value(resumen)
		sheet.cell(column = Columnas.CResumen.value, row = fila).font = ft_normal
		#precio
		sheet.cell(column = Columnas.CPrecio.value, row = fila).value = consulta.value(precio)
		sheet.cell(column = Columnas.CPrecio.value, row = fila).font = ft_normal
		sheet.cell(column = Columnas.CPrecio.value, row = fila).number_format = '#,##0.00'
		#importe
		coordenadaCantidad= get_column_letter(Columnas.CCantidad.value)+str(fila)
		coordenadaPrecio= get_column_letter(Columnas.CPrecio.value)+str(fila)
		sheet.cell(column = Columnas.CImporte.value, row = fila).font = ft_normal
		sheet.cell(column = Columnas.CImporte.value, row = fila).number_format = '#,##0.00'
		coordenadaImporte= get_column_letter(Columnas.CImporte.value)+str(fila)
		sheet[coordenadaImporte]= "="+coordenadaCantidad+"*"+coordenadaPrecio
		fila = fila +1
	#anchos
	ancho_columnas = 0
	'''for column_cells in sheet.columns:
		ancho = max(len(as_text(cell.value)) for cell in column_cells)
		sheet.column_dimensions[column_cells[0].column].width = ancho
		ancho_columnas = '''
	#columna del codigo
	columnaCodigo = sheet[get_column_letter(Columnas.CCodigo.value)]
	for cell in columnaCodigo:
		ancho = max(len(as_text(cell.value)) for cell in columnaCodigo)
		sheet.column_dimensions[get_column_letter(Columnas.CCodigo.value)].width = ancho
	ancho_columnas +=ancho
	#columna cantidad
	columnaCantidad = sheet[get_column_letter(Columnas.CCantidad.value)]
	for cell in columnaCantidad:
		ancho = max(len(as_text(cell.value)) for cell in columnaCantidad)
		sheet.column_dimensions[get_column_letter(Columnas.CCantidad.value)].width = ancho
	ancho_columnas +=ancho
	#columna ud
	columnaUnidad = sheet[get_column_letter(Columnas.CUd.value)]
	for cell in columnaUnidad:
		ancho = max(len(as_text(cell.value)) for cell in columnaUnidad)
		sheet.column_dimensions[get_column_letter(Columnas.CUd.value)].width = ancho+1
	ancho_columnas +=ancho
	#columna precio
	columnaPrecio = sheet[get_column_letter(Columnas.CPrecio.value)]
	for cell in columnaPrecio:
		ancho = max(len(as_text(cell.value)) for cell in columnaPrecio)
		sheet.column_dimensions[get_column_letter(Columnas.CPrecio.value)].width = ancho
	ancho_columnas +=ancho
	#columna importe
	columnaImporte = sheet[get_column_letter(Columnas.CImporte.value)]
	for cell in columnaImporte:
		ancho = max(len(as_text(cell.value)) for cell in columnaImporte)
		sheet.column_dimensions[get_column_letter(Columnas.CImporte.value)].width = ancho
	ancho_columnas +=ancho
	#columna resumen (ajusto esta columna al final porque su ancho sera el ancho de la pagina menos los anchos ya ajustados -ancho_columnas-)
	columnaResumen = sheet[get_column_letter(Columnas.CResumen.value)]
	#for cell in columnaResumen:
	#	ancho = max(len(as_text(cell.value)) for cell in columnaResumen)
	sheet.column_dimensions[get_column_letter(Columnas.CResumen.value)].width = 65-ancho_columnas
	
	
		
		
	#ajuste del texto a la pagina
	#sheet.sheet_properties.pageSetUpPr.fitToPage = True
			
	book.save('listado_conceptos.xlsx')
	
def as_text(value):
    if value is None:
        return ""
    return str(value)
