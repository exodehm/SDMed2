from PyQt5 import QtSql
from openpyxl import Workbook
from openpyxl.worksheet import dimensions
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Font
from enum import Enum

import datetime

class Columnas(Enum):
    CCodigo = 1
    CResumen = 2
    CUds = 3
    CLongitud = 4
    CAnchura = 5
    CAltura = 6
    CParciales = 7
    CCantidad = 8
    CPrecio = 9
    CImporte = 10

def run():
    print("Plugin para imprimir mediciones y presupuestos")
    
def imprimir(conexion, obra, book):
	tipoInforme = "Presupuestos y Mediciones"
	imprimirMediciones = False
	#abro una instancia de hoja de calculo
	sheet = book.active
	book.remove_sheet(sheet)
	#fuentes
	ft_resaltada = Font(name='Arial', size=11, bold=True)
	ft_normal = Font(name='Arial', size=10, bold=False)
	ft_negrita = Font(name='Arial', size=10, bold=True)
	ft_chica = Font(name='Arial', size=8, bold=False)		
	#cabecera
	consulta = QtSql.QSqlQuery(conexion)
	subconsulta = QtSql.QSqlQuery(conexion)
	consultamediciones = QtSql.QSqlQuery(conexion)
	consulta.exec_("SELECT resumen FROM \"" + obra + "_Conceptos\" AS C, \"" + obra + "_Relacion\" AS R WHERE C.codigo = R.codhijo AND R.codpadre IS NULL")
	print (consulta.lastQuery())
	print (consulta.lastError().text())
	resumen = ""
	while consulta.next():
		resumen = consulta.value(0)
	#definirCabecera	
	#e inserto datos
	fila = 1
	#capitulos
	consulta.exec_("SELECT C.codigo, C.resumen FROM \"" + obra + "_Conceptos\" AS C, \"" + obra + "_Relacion\" AS R WHERE R.codpadre ='"+obra+"' AND C.codigo = R.codhijo")
	print (consulta.lastQuery())
	print (consulta.lastError().text())
	codigo = ""
	resumen = ""	
	while consulta.next():		
		codigo = consulta.value(0)
		resumen = consulta.value(1)
		sheet = book.create_sheet(title=resumen)
		sheet.merge_cells(start_row=fila, start_column=Columnas.CCodigo.value, end_row=fila, end_column=Columnas.CImporte.value)
		sheet.cell(column = Columnas.CCodigo.value, row = fila).value = codigo + " " + resumen
		fila = fila + 1;
		#partidas
		subconsulta.exec_("SELECT C.codigo, C.ud, C.resumen , R.canpres, C.preciomed, C.descripcion FROM \"" + obra + "_Conceptos\" AS C, \"" + obra + "_Relacion\" AS R WHERE R.codpadre ='"+codigo+"' AND C.codigo = R.codhijo")
		codigopartida=""
		udpartida = ""
		resumenpartida = ""
		descripcionpartida=""
		cantidad = 0
		precio = 0
		while subconsulta.next():
			codigopartida = subconsulta.value(0)
			udpartida = subconsulta.value(1)
			resumenpartida = subconsulta.value(2)
			cantidad = float(subconsulta.value(3))
			precio = float(subconsulta.value(4))
			descripcionpartida = subconsulta.value(5)			
			#codigo partida
			sheet.cell(column = Columnas.CCodigo.value, row = fila).value = codigopartida
			sheet.cell(column = Columnas.CCodigo.value, row = fila).font = ft_negrita
			#ud y resumen de partida (juntos)
			sheet.merge_cells(start_row=fila, start_column=Columnas.CResumen.value, end_row=fila, end_column=Columnas.CPrecio.value)
			sheet.cell(column = Columnas.CResumen.value, row = fila).value = udpartida + " " + resumenpartida
			sheet.cell(column = Columnas.CResumen.value, row = fila).font = ft_negrita
			fila = fila + 1;
			#descripcion
			anchocolumna = getApproximateArialStringWidth(descripcionpartida)
			#print ("ancho columna" + str(anchocolumna))
			num_lineas = anchocolumna/13
			sheet.merge_cells(start_row=fila, start_column=Columnas.CResumen.value, end_row=fila, end_column=Columnas.CPrecio.value)
			sheet.cell(column = Columnas.CResumen.value, row = fila).value = descripcionpartida
			sheet.cell(column = Columnas.CResumen.value, row = fila).alignment = Alignment(wrapText=True, vertical="top")			
			sheet.cell(column = Columnas.CResumen.value, row = fila).font = ft_normal
			sheet.row_dimensions[fila].height = num_lineas * ft_normal.size	* 0.6
			#sheet.cell(column = Columnas.CResumen.value, row = fila).alignment = Alignment()
			fila = fila + 1;
			#mediciones (opcional)
			if imprimirMediciones == True:
				tipo = 0
				comentario = ""
				uds= 0
				longitud = 0
				anchura = 0
				altura = 0
				fila_inicial = fila
				consultamediciones.exec_("SELECT * FROM ver_lineas_medicion('"+ obra + "','" + codigo +"','"+codigopartida+"')")
				#print (consultamediciones.lastQuery()) 
				while consultamediciones.next():
					tipo = consultamediciones.value(0)
					comentario = consultamediciones.value(1)
					uds = consultamediciones.value(2)
					longitud = consultamediciones.value(3)
					anchura = consultamediciones.value(4)
					altura = consultamediciones.value(5)
					#comentario
					sheet.cell(column = Columnas.CResumen.value, row = fila).value = comentario
					sheet.cell(column = Columnas.CResumen.value, row = fila).font = ft_chica
					#uds					
					escribirNumero(sheet,fila,Columnas.CUds.value,uds,ft_chica)
					coordenadaUds= get_column_letter(Columnas.CUds.value)+str(fila)
					#longitud
					escribirNumero(sheet,fila,Columnas.CLongitud.value,longitud,ft_chica)
					coordenadaLong= get_column_letter(Columnas.CLongitud.value)+str(fila)
					#anchura
					escribirNumero(sheet,fila,Columnas.CAnchura.value,anchura,ft_chica)
					coordenadaLat= get_column_letter(Columnas.CAnchura.value)+str(fila)
					#altura
					escribirNumero(sheet,fila,Columnas.CAltura.value,altura,ft_chica)
					coordenadaAlt= get_column_letter(Columnas.CAltura.value)+str(fila)
					#parcial
					coordenadaParcial= get_column_letter(Columnas.CParciales.value)+str(fila)
					sheet[coordenadaParcial]="= IF("+coordenadaUds+"<>0,"+coordenadaUds+",1) * IF("+coordenadaLong+"<>0,"+coordenadaLong+",1) * IF("+coordenadaLat+"<>0,"+coordenadaLat+",1) * IF("+coordenadaAlt+"<>0,"+coordenadaAlt+",1)"
					sheet[coordenadaParcial].font = ft_chica
					sheet[coordenadaParcial].number_format = '#,##0.00'
					fila = fila +1
				#linea de suma
				#sheet.cell(column = Columnas.CParciales.value, row = fila).border = Border(top=Side(style='thin'))#linea superior
				#sheet.cell(column = Columnas.CCantidad.value, row = fila).border = Border(top=Side(style='thin'))#linea superior
				#sheet.cell(column = Columnas.CPrecio.value, row = fila).border = Border(top=Side(style='thin'))#linea superior
				#sheet.cell(column = Columnas.CImporte.value, row = fila).border = Border(top=Side(style='thin'))#linea superior	
				#cantidad			
				#cadena_rango = get_column_letter(Columnas.CParciales.value)+str(fila_inicial)+":"+get_column_letter(Columnas.CParciales.value)+str(fila-1)
				#coordenadaSumatorio = get_column_letter(Columnas.CCantidad.value)+str(fila)
				#if imprimirMediciones is True:
				#	sheet[coordenadaSumatorio]="=SUM("+cadena_rango+")"
				#else:
				sheet[coordenadaSumatorio]=cantidad
				sheet[coordenadaSumatorio].font = ft_chica
				sheet[coordenadaSumatorio].number_format = '#,##0.00'
				#precio
				escribirNumero(sheet,fila,Columnas.CPrecio.value,precio,ft_chica)
				coordenadaPrecio= get_column_letter(Columnas.CPrecio.value)+str(fila)
				#importe
				coordenadaImporte= get_column_letter(Columnas.CImporte.value)+str(fila)				
				sheet[coordenadaImporte]="="+coordenadaSumatorio+"*"+coordenadaPrecio
				sheet.cell(column = Columnas.CImporte.value, row = fila).number_format = '#,##0.00'			
				sheet.cell(column = Columnas.CImporte.value, row = fila).font = ft_chica
				fila = fila +3
		sheet.cell(column=Columnas.CResumen.value, row = fila).value = "TOTAL CAPITULO KK" + codigo + " " + resumen
		sheet.cell(column=Columnas.CResumen.value, row = fila).font = ft_negrita
		#cantidad total del capitulo
		cadena_rango = get_column_letter(Columnas.CImporte.value)+str(1)+":"+get_column_letter(Columnas.CImporte.value)+str(fila-1)
		coordenadaTotalCapitulo = get_column_letter(Columnas.CImporte.value)+str(fila)
		sheet[coordenadaTotalCapitulo]="=SUM("+cadena_rango+")"
		sheet[coordenadaTotalCapitulo].font = ft_negrita
		sheet[coordenadaTotalCapitulo].number_format = '#,##0.00'
		fila = 1
		#fijo el ancho de la columna de resumen
		sheet.column_dimensions[get_column_letter(Columnas.CResumen.value)].width = 15
		#fijo ancho de columnas numericas (uds,longitud, anchura,altura,parcial,cantidad,precio,importe
		for i in range (Columnas.CUds.value, Columnas.CImporte.value):
			columnaAjustar = sheet[get_column_letter(i)]
			#ancho = max(len(as_text(cell.value)) for cell in columnaAjustar)		
			sheet.column_dimensions[get_column_letter(i)].width = 5
		#defino el pie de pagina
		definirPie(sheet)
	
	book.save('presupuestos.xlsx')
	
def as_text(value):
    if value is None:
        return ""
    return str(value)
    
def getApproximateArialStringWidth(st):
	size = 0 # in milinches
	for s in st:
		if s in 'lij|\' ':
			size += 37
		elif s in '![]fI.,:;/\\t':
			size += 50
		elif s in '`-(){}r"':
			size += 60
		elif s in '*^zcsJkvxy':
			size += 85
		elif s in 'aebdhnopqug#$L+<>=?_~FZT0123456789':
			size += 95
		elif s in 'BSPEAKVXY&UwNRCHD':
			size += 112
		elif s in 'QGOMm%W@':
			size += 135
		else: size += 50
	return size * 6 / 1000.0 # Convert to picas
	
def escribirNumero (sheet,fila,columna,valor,fuente):
	if valor == 0:
		sheet.cell(column = columna, row = fila).value = ""
	else:
		sheet.cell(column = columna, row = fila).value = valor
		sheet.cell(column = columna, row = fila).font = fuente
		sheet.cell(column = columna, row = fila).number_format = '#,##0.00'
		
def definirCabecera(sheet, resumen):
	sheet.oddHeader.left.size = 10
	sheet.oddHeader.left.font = "Arial,Bold"
	sheet.oddHeader.left.text = tipoInforme.upper() + "\n" +resumen
	sheet.oddHeader.left.size = 9
	sheet.oddHeader.left.font = "Arial"
	sheet.oddHeader.left.text = "\n Codigo         Resumen\n______________________________"
	sheet.oddHeader.center.size = 10
	sheet.oddHeader.center.font = "Arial,Bold"
	sheet.oddHeader.right.size = 10
	sheet.oddHeader.right.font = "Arial,Bold"	
	sheet.oddHeader.center.text = "\n\nUds   Longitud   Anchura    Altura\n______________________________"
	sheet.oddHeader.right.text = "\n\nParciales  Cantidad Precio Importe\n______________________________"
	sheet.print_title_cols = 'A:H'
	sheet.print_title_rows = '1:3'	
	
def definirPie(sheet):
	sheet.oddFooter.center.size = 11
	sheet.oddFooter.center.font = "Arial,Bold"	
	sheet.oddFooter.center.text = "Página &[Page] de &N"	
	
	
