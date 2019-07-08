from PyQt5 import QtSql
from openpyxl import Workbook
from openpyxl.drawing import line
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Font
import datetime

def run():
    print("Plugin para imprimir resumen de presupuesto")
    
def imprimir(conexion, obra, book):	
	#abro una instancia de hoja de calculo	
	sheet = book.active
	sheet.title = "RESUMEN DE PRESUPUESTO"
	#cabecera
	consulta = QtSql.QSqlQuery("SELECT resumen FROM \"" + obra + "_Conceptos\" AS C, \"" + obra + "_Relacion\" AS R WHERE C.codigo = R.codhijo AND R.codpadre IS NULL", conexion)
	print (consulta.lastQuery())
	print (consulta.lastError().text())
	resumen = ""
	while consulta.next():
		resumen = consulta.value(0)		
	sheet.oddHeader.left.size = 11
	sheet.oddHeader.left.font = "Arial,Bold"
	#sheet.oddHeader.left.color = "CC3366"
	sheet.oddHeader.left.text = "RESUMEN DE PRESUPUESTO " + resumen	
	sheet.print_title_cols = 'A:C'
	sheet.print_title_rows = '1:2'
	#pie
	sheet.oddFooter.center.size = 11
	sheet.oddFooter.center.font = "Arial,Bold"	
	sheet.oddFooter.center.text = "Página &[Page] de &N"	
	#datos para situar filas y columnas
	fila_inicial = 0
	fila_final = 0
	fila = 1	
	offset_izquierdo = 1
	columna = offset_izquierdo
	columna_euros=5
	letra_columna_euros = get_column_letter(columna_euros)
	contador = 0
	cantidad = 0
	ancho_columnas = 0
	#fuentes
	ft_resaltada = Font(name='Arial', size=11, bold=True)
	ft_normal = Font(name='Arial', size=10, bold=False)
	ft_chica = Font(name='Arial', size=9, bold=False)
	#cabecera
	consulta.exec_("SELECT codigo,resumen,canpres*preciomed AS \"EUROS\" FROM \"" + obra + "_Conceptos\" AS C, \"" + obra + "_Relacion\" AS R WHERE C.codigo = R.codhijo AND R.codpadre = '" + obra + "' ORDER BY R.posicion")
	print (consulta.lastError().text())
	rec = consulta.record()
	codigo = rec.indexOf("codigo")
	resumen = rec.indexOf("resumen")
	euros = rec.indexOf("EUROS")
	sheet.cell(column = columna, row = fila).value = "CAPITULO"
	sheet.cell(column = columna, row = fila).font = ft_normal
	sheet.cell(column = columna+1, row = fila).value = "RESUMEN"
	sheet.cell(column = columna+1, row = fila).font = ft_normal
	sheet.cell(column = columna_euros, row = fila).value = "EUROS"
	sheet.cell(column = columna_euros, row = fila).font = ft_normal
	sheet.cell(column = columna_euros, row = fila).alignment = Alignment(horizontal="right", vertical="center")
	sheet.cell(column = columna_euros+1, row = fila).value = "%"
	sheet.cell(column = columna_euros+1, row = fila).font = ft_normal
	sheet.cell(column = columna_euros+1, row = fila).alignment = Alignment(horizontal="right", vertical="center")
	#linea
	for i in range(offset_izquierdo,columna_euros+2):
		sheet.cell(column = i, row = fila).border = Border(bottom=Side(style='thick'))		
	#datos
	fila = fila +1
	while consulta.next():				
		#codigo
		sheet.cell(column = columna, row = fila).value = consulta.value(codigo)
		sheet.cell(column = columna, row = fila).font = ft_normal
		#resumen
		columna =  columna + 1
		sheet.cell(column = columna, row = fila).value = consulta.value(resumen)
		sheet.cell(column = columna, row = fila).font = ft_normal
		#euros
		columna =  columna_euros
		sheet.cell(column = columna, row = fila).value = consulta.value(euros)
		sheet.cell(column = columna, row = fila).font = ft_normal
		sheet.cell(column = columna, row = fila).number_format = '#,##0.00'
		cantidad = cantidad +consulta.value(euros)
				
		if contador == 0:
			fila_inicial = fila		
		if contador == consulta.size()-1:
			fila_final = fila		
		fila = fila + 1
		contador = contador +1
		columna = offset_izquierdo
	#porcentajes (a la derecha de las cantidades de cada capitulo)
	for i in range(0,consulta.size()):
		#sheet.cell(column = columna_euros+1, row = fila_inicial+i).value = (sheet.cell(row=fila_inicial+i, column=columna_euros).value)/cantidad
		sheet.cell(column = columna_euros+1, row = fila_inicial+i).value = "=ROUND("+str((sheet.cell(row=fila_inicial+i, column=columna_euros).value)/cantidad)+",4)"
		sheet.cell(column = columna_euros+1, row = fila_inicial+i).font = ft_chica
		sheet.cell(column = columna_euros+1, row = fila_inicial+i).number_format = '0.00%'
	#unir celdas entre resumen y precio y rellenar de punteado
	for i in range(0,consulta.size()):
		sheet.merge_cells(start_row=fila_inicial+i, start_column=offset_izquierdo+2, end_row=fila_inicial+i, end_column=columna_euros-1)		
		coordenada_punteado_CAP = get_column_letter(offset_izquierdo+2)+str(fila_inicial+i)#punteado
		sheet[coordenada_punteado_CAP].border = Border(bottom=Side(style='dotted'))#punteado		
	#antes de seguir ajustamos el ancho	de las columnas		
	for column_cells in sheet.columns:
		ancho = max(len(as_text(cell.value)) for cell in column_cells)
		sheet.column_dimensions[column_cells[0].column].width = ancho
		ancho_columnas = ancho_columnas + ancho
	sheet.column_dimensions['F'].width = 8
	#EM
	fila = fila + 1;
	sheet.merge_cells(start_row=fila, start_column=offset_izquierdo, end_row=fila, end_column=columna_euros-1)
	coordenada_leyenda_EM = get_column_letter(1)+str(fila)
	celda_leyenda_Ejecucion_Material =sheet[coordenada_leyenda_EM]
	celda_leyenda_Ejecucion_Material.value = "Total Ejecución Material: "
	celda_leyenda_Ejecucion_Material.alignment = Alignment(horizontal="right", vertical="center")
	celda_leyenda_Ejecucion_Material.font = ft_resaltada #negritas	
	coordenadaEM = letra_columna_euros+str(fila)
	cadena_rango = letra_columna_euros+str(fila_inicial)+":"+letra_columna_euros+str(fila_final)	
	sheet[coordenadaEM]="=SUM("+cadena_rango+")"
	sheet[coordenadaEM].font = ft_resaltada #negritas
	sheet[coordenadaEM].border = Border(top=Side(style='thin'))#linea superior 
	sheet[coordenadaEM].number_format = '#,##0.00'	
	#GG
	#consulta para hallar el dato GG
	gastos_generales = 0.0
	consulta.exec_("SELECT datos->>'Valor' FROM (SELECT jsonb_array_elements(propiedades->'Valor') AS datos \
				FROM \"" + obra + "_Propiedades\" \
				WHERE propiedades->>'Propiedad' = 'Porcentajes') AS subdatos WHERE datos->>'Variable' = 'zPorGastosGenerales'")	
	while consulta.next():
		gastos_generales = float(consulta.value(0))	
	fila = fila +1
	sheet.merge_cells(start_row=fila, start_column=offset_izquierdo, end_row=fila, end_column=columna_euros-3)
	coordenada_leyenda_GG = get_column_letter(1)+str(fila)
	celda_leyenda_Gastos_Generales =sheet[coordenada_leyenda_GG]
	celda_leyenda_Gastos_Generales.value = str(gastos_generales)+"% Gastos Generales: "
	celda_leyenda_Gastos_Generales.alignment = Alignment(horizontal="right", vertical="center")
	celda_leyenda_Gastos_Generales.font = ft_normal
	coordenadaGG = get_column_letter(columna_euros-1)+str(fila)
	sheet[coordenadaGG]="="+coordenadaEM+"*"+str(gastos_generales)+"/"+str(100)
	sheet[coordenadaGG].font = ft_normal
	sheet[coordenadaGG].number_format = '#,##0.00'
	coordenada_punteado_GG = get_column_letter(columna_euros-2)+str(fila)#punteado
	sheet[coordenada_punteado_GG].border = Border(bottom=Side(style='dotted'))#punteado	
	#BI
	#consulta para hallar el dato BI
	beneficio_industrial = 0.0
	consulta.exec_("SELECT datos->>'Valor' FROM (SELECT jsonb_array_elements(propiedades->'Valor') AS datos \
				FROM \"" + obra + "_Propiedades\" \
				WHERE propiedades->>'Propiedad' = 'Porcentajes') AS subdatos WHERE datos->>'Variable' = 'zPorBenIndustrial'")	
	while consulta.next():
		beneficio_industrial = float(consulta.value(0))	
	fila = fila +1
	sheet.merge_cells(start_row=fila, start_column=offset_izquierdo, end_row=fila, end_column=columna_euros-3)
	coordenada_leyenda_BI = get_column_letter(1)+str(fila)
	celda_leyenda_Beneficio_Industrial =sheet[coordenada_leyenda_BI]
	celda_leyenda_Beneficio_Industrial.value = str(beneficio_industrial)+"% Beneficio Industrial: "
	celda_leyenda_Beneficio_Industrial.alignment = Alignment(horizontal="right", vertical="center")
	celda_leyenda_Beneficio_Industrial.font = ft_normal
	coordenadaBI= get_column_letter(columna_euros-1)+str(fila)
	sheet[coordenadaBI]="="+coordenadaEM+"*"+str(beneficio_industrial)+"/"+str(100)
	sheet[coordenadaBI].font = ft_normal
	sheet[coordenadaBI].number_format = '#,##0.00'
	coordenada_punteado_BI = get_column_letter(columna_euros-2)+str(fila)#punteado
	sheet[coordenada_punteado_BI].border = Border(bottom=Side(style='dotted'))#punteado	
	
	cantidad = cantidad + cantidad*(gastos_generales+beneficio_industrial)/100
	#suma de GG+BI
	fila = fila +1
	sheet.merge_cells(start_row=fila, start_column=offset_izquierdo, end_row=fila, end_column=columna_euros-1)
	sheet.cell(column = 1,row = fila, value = "SUMA DE G.G. y B.I.: ")
	sheet.cell(column = 1,row = fila).alignment = Alignment(horizontal="right", vertical="center")
	sheet.cell(column = 1,row = fila).font = ft_normal
	sheet.cell(column = 1,row = fila).number_format = '#,##0.00'
	coordenadaGGMASBI= get_column_letter(columna_euros)+str(fila)
	sheet[coordenadaGGMASBI]="="+coordenadaGG+"+"+coordenadaBI
	sheet[coordenadaGGMASBI].border = Border(top=Side(style='thin'))#linea superior 
	sheet[coordenadaGGMASBI].font = ft_normal
	sheet[coordenadaGGMASBI].number_format = '#,##0.00'
	#presupuesto de contrata
	fila = fila +1
	sheet.merge_cells(start_row=fila, start_column=offset_izquierdo, end_row=fila, end_column=columna_euros-1)
	coordenada_leyenda_PC = get_column_letter(1)+str(fila)
	celda_leyenda_Presupuesto_Contrata =sheet[coordenada_leyenda_PC]
	celda_leyenda_Presupuesto_Contrata.value = "TOTAL PRESUPUESTO DE CONTRATA"
	celda_leyenda_Presupuesto_Contrata.alignment = Alignment(horizontal="right", vertical="center")
	celda_leyenda_Presupuesto_Contrata.font = ft_resaltada
	coordenadaPC= get_column_letter(columna_euros)+str(fila)
	sheet[coordenadaPC]="="+coordenadaEM+"+"+coordenadaGGMASBI
	sheet[coordenadaPC].border = Border(top=Side(style='thin'))#linea superior 
	sheet[coordenadaPC].font = ft_resaltada #negritas
	sheet[coordenadaPC].number_format = '#,##0.00'
	#iva
	#consulta para hallar el dato IVA
	IVA = 0.0
	consulta.exec_("SELECT datos->>'Valor' FROM (SELECT jsonb_array_elements(propiedades->'Valor') AS datos \
				FROM \"" + obra + "_Propiedades\" \
				WHERE propiedades->>'Propiedad' = 'Porcentajes') AS subdatos WHERE datos->>'Variable' = 'zPorIVAEjecucion'")	
	while consulta.next():
		IVA = float(consulta.value(0))	
	fila = fila +1
	sheet.merge_cells(start_row=fila, start_column=offset_izquierdo, end_row=fila, end_column=columna_euros-3)
	coordenada_leyenda_IVA = get_column_letter(1)+str(fila)
	celda_leyenda_IVA =sheet[coordenada_leyenda_IVA]
	celda_leyenda_IVA.value = str(IVA) + "% IVA"
	celda_leyenda_IVA.alignment = Alignment(horizontal="right", vertical="center")
	celda_leyenda_IVA.font = ft_normal
	coordenadaIVA= get_column_letter(columna_euros-1)+str(fila)
	sheet[coordenadaIVA]="="+coordenadaPC+"*"+str(IVA)+"/"+str(100)
	sheet[coordenadaIVA].font = ft_normal
	sheet[coordenadaIVA].number_format = '#,##0.00'
	coordenada_punteado_IVA = get_column_letter(columna_euros-2)+str(fila)#punteado
	sheet[coordenada_punteado_IVA].border = Border(bottom=Side(style='dotted'))#punteado	
	
	cantidad = cantidad + cantidad*IVA/100
	#total presupuesto
	fila = fila +1
	sheet.merge_cells(start_row=fila, start_column=offset_izquierdo, end_row=fila, end_column=columna_euros-1)
	coordenada_leyenda_TP = get_column_letter(1)+str(fila)
	celda_leyenda__Total_Presupuesto =sheet[coordenada_leyenda_TP]
	celda_leyenda__Total_Presupuesto.value = "TOTAL PRESUPUESTO GENERAL"
	celda_leyenda__Total_Presupuesto.alignment = Alignment(horizontal="right", vertical="center")
	celda_leyenda__Total_Presupuesto.font = ft_resaltada
	coordenadaTP= get_column_letter(columna_euros)+str(fila)
	sheet[coordenadaTP]="="+coordenadaPC+"+"+coordenadaIVA
	sheet[coordenadaTP].border = Border(top=Side(style='thin'))#linea superior
	sheet[coordenadaTP].font = ft_resaltada #negritas
	sheet[coordenadaTP].number_format = '#,##0.00'
	
	#termino de ajuStar los anchos 	
	columnaD = sheet['D']
	ancho = max(len(as_text(cell.value)) for cell in columnaD)
	sheet.column_dimensions['D'].width = ancho
	ancho_columnas = ancho_columnas + ancho
	print (ancho_columnas)
	sheet.column_dimensions['C'].width = 90-ancho_columnas
	#cd = sheet['D']
	#print (cd[0])
		
	#linea texto con cantidad final
	print ("cantidad "+ str(cantidad))
	consulta.exec_("SELECT numero_en_euro("+str(cantidad)+")")
	print (consulta.lastError().text())
	cantidadenletra=""
	while consulta.next():
		cantidadenletra = consulta.value(0)
	print ("cantidad en letras "+cantidadenletra)
	fila = fila +3
	sheet.merge_cells(start_row=fila, start_column=offset_izquierdo, end_row=fila, end_column=columna_euros+1)
	sheet.cell(column = offset_izquierdo,row = fila, value = "Asciende el presupuesto general a la expresada cantidad de " + str(cantidadenletra))	
	sheet.cell(column = offset_izquierdo,row = fila).font = ft_normal
	sheet.cell(column = offset_izquierdo,row = fila).alignment = Alignment(wrap_text = True)
	rd = sheet.row_dimensions[fila]
	rd.height = 25 # value in points, there is no "auto"
	
	#ciudad y fecha
	fila = fila +3
	sheet.merge_cells(start_row=fila, start_column=offset_izquierdo, end_row=fila, end_column=columna_euros)
	ciudad = "Granada"
	hoy = datetime.date.today()	
	sheet.cell(column = offset_izquierdo,row = fila, value = ciudad + ", a " + str(hoy.strftime('%d de %b de %Y')))
	sheet.cell(column = offset_izquierdo,row = fila).alignment = Alignment(horizontal="center", vertical="center")
	sheet.cell(column = offset_izquierdo,row = fila).font = ft_normal
	
	#firmas-datos
	fila = fila + 4
	encabezado_firma_proyectista = ""
	nombre_proyectista1 = ""
	nombre_proyectista2 = ""
	encabezado_firma_promotor = ""
	nombre_promotor1 = ""
	nombre_promotor2 = ""
	consulta.exec_("SELECT datos->>'Valor' FROM (SELECT jsonb_array_elements(propiedades->'Valor') AS datos \
				FROM \"" + obra + "_Propiedades\" \
				WHERE propiedades->>'Propiedad' = 'Proyectista') AS subdatos WHERE datos->>'Variable' = 'zPryEncabezamiento'")
	while consulta.next():
		encabezado_firma_proyectista = consulta.value(0)
	consulta.exec_("SELECT datos->>'Valor' FROM (SELECT jsonb_array_elements(propiedades->'Valor') AS datos \
				FROM \"" + obra + "_Propiedades\" \
				WHERE propiedades->>'Propiedad' = 'Proyectista') AS subdatos WHERE datos->>'Variable' = 'zPryNombre1'")
	while consulta.next():
		nombre_proyectista1 = consulta.value(0)
	consulta.exec_("SELECT datos->>'Valor' FROM (SELECT jsonb_array_elements(propiedades->'Valor') AS datos \
				FROM \"" + obra + "_Propiedades\" \
				WHERE propiedades->>'Propiedad' = 'Proyectista') AS subdatos WHERE datos->>'Variable' = 'zPryNombre2'")
	while consulta.next():
		nombre_proyectista2 = consulta.value(0)
	consulta.exec_("SELECT datos->>'Valor' FROM (SELECT jsonb_array_elements(propiedades->'Valor') AS datos \
				FROM \"" + obra + "_Propiedades\" \
				WHERE propiedades->>'Propiedad' = 'El promotor') AS subdatos WHERE datos->>'Variable' = 'zProEncabezamiento'")
	while consulta.next():
		encabezado_firma_promotor = consulta.value(0)
	consulta.exec_("SELECT datos->>'Valor' FROM (SELECT jsonb_array_elements(propiedades->'Valor') AS datos \
				FROM \"" + obra + "_Propiedades\" \
				WHERE propiedades->>'Propiedad' = 'El promotor') AS subdatos WHERE datos->>'Variable' = 'zProNombre1'")
	while consulta.next():
		nombre_promotor1 = consulta.value(0)
	consulta.exec_("SELECT datos->>'Valor' FROM (SELECT jsonb_array_elements(propiedades->'Valor') AS datos \
				FROM \"" + obra + "_Propiedades\" \
				WHERE propiedades->>'Propiedad' = 'El promotor') AS subdatos WHERE datos->>'Variable' = 'zProNombre2'")
	while consulta.next():
		nombre_promotor2 = consulta.value(0)
	#firmas-situar
	sheet.cell(column = offset_izquierdo,row = fila).font = ft_resaltada
	sheet.cell(column = offset_izquierdo,row = fila, value = encabezado_firma_promotor)	
	sheet.cell(column = offset_izquierdo+4,row = fila).font = ft_resaltada
	sheet.cell(column = offset_izquierdo+4,row = fila, value = encabezado_firma_proyectista)
	fila = fila + 3
	sheet.cell(column = offset_izquierdo,row = fila).font = ft_normal
	sheet.cell(column = offset_izquierdo,row = fila, value = nombre_promotor1)		
	sheet.cell(column = offset_izquierdo+4,row = fila).font = ft_normal
	sheet.cell(column = offset_izquierdo+4,row = fila, value = nombre_proyectista1)
	fila = fila + 1
	sheet.cell(column = offset_izquierdo,row = fila).font = ft_normal
	sheet.cell(column = offset_izquierdo,row = fila, value = nombre_promotor2)	
	sheet.cell(column = offset_izquierdo+4,row = fila).font = ft_normal
	sheet.cell(column = offset_izquierdo+4,row = fila, value = nombre_proyectista2)	
	#ajuste del texto a la pagina
	#sheet.sheet_properties.pageSetUpPr.fitToPage = True
	
	book.save('resumen.xlsx')
	
def as_text(value):
    if value is None:
        return ""
    return str(value)
	
	
